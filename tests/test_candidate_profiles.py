from pathlib import Path

from openpyxl import load_workbook
import pytest

from mordheim_optimizer.candidate_catalog import (
    armour_descriptions,
    equipment_costs_for_profile,
    equipment_options_for_profile,
    find_profile,
    load_bands,
    weapon_descriptions,
)
from mordheim_optimizer.ui import MordheimApp
from mordheim_optimizer.rules import BODY_ARMORS, SKILLS, WEAPONS_ALL
from mordheim_optimizer.workbooks import (
    CandidateWorkbookError,
    DATA_SHEET,
    ENEMIES_SHEET,
    FORMAT_VERSION,
    RESULTS_INDEX_SHEET,
    RESULT_SHEET_PREFIX,
    SUMMARY_SHEET,
    load_candidate_workbook,
    save_candidate_workbook,
)


def _payload():
    return {
        "config": {
            "candidate_name": "Rat with papers",
            "candidate_band_id": "mordheim-skaven",
            "candidate_profile_id": "eshin-assassin",
            "HA": 4, "F": 4, "R": 3, "H": 1, "I": 5, "A": 1,
            "skills": ["Mighty Blow"], "main_weapon": "Sword",
            "off_hand": "Dagger", "main_weapon_material": "Normal",
            "offhand_material": "Normal", "armor": "Light armour",
            "has_helmet": True,
        },
        "candidate": {
            "name": "Rat with papers", "band_name": "Skaven Clan Eshin",
            "profile_name": "Asesino", "profile_type": "hero",
            "rules": ["Luchador Consumado: -1 additional a the save."],
        },
        "opponent": {"mode": "Random sample", "level": 2, "description": "Low, Medium"},
        "enemies": {
            "mode": "custom", "level": 2, "difficulties": [],
            "profiles": [
                {
                    "enemy_name": "Brute", "HA": 3, "F": 4, "R": 4,
                    "H": 1, "I": 2, "A": 1, "skills": [],
                    "main_weapon": "Mace", "off_hand": "None",
                    "main_weapon_material": "Normal",
                    "offhand_material": "Normal", "armor": "No Armour",
                    "has_helmet": False,
                },
                {
                    "enemy_name": "Fast", "HA": 4, "F": 3, "R": 3,
                    "H": 1, "I": 5, "A": 2, "skills": ["Lightning Reflexes"],
                    "main_weapon": "Sword", "off_hand": "Dagger",
                    "main_weapon_material": "Normal",
                    "offhand_material": "Normal", "armor": "Light armour",
                    "has_helmet": True,
                },
            ],
        },
    }


def test_equipment_filter_respects_profile_specific_restrictions():
    beastmaster = set(equipment_options_for_profile("dark-elves", "beastmaster"))
    corsair = set(equipment_options_for_profile("dark-elves", "corsairs"))
    assert {"Black Venom", "Sea Dragon cloak", "Light armour"} <= beastmaster
    assert "Lucky charm" in beastmaster
    assert "Lucky charm" not in corsair


def test_general_market_equipment_respects_faction_restrictions():
    matriarch = set(equipment_options_for_profile("sisters-of-sigmar", "sigmarite-matriarch"))
    assassin = set(equipment_options_for_profile("skaven-clan-eshin", "assassin-adept"))
    assert "Lucky charm" in matriarch
    assert "Spider Spittle" in assassin
    assert "Black Lotus" not in matriarch
    assert "Black Venom" not in matriarch


def test_equipment_costs_use_band_prices_and_expected_dice_values():
    costs = equipment_costs_for_profile(
        "dark-elves", "beastmaster"
    )
    assert costs["Light armour"] == 20.0
    assert costs["Sea Dragon cloak"] == 50.0
    assert equipment_costs_for_profile()["Lucky charm"] == 10.0


def test_sea_dragon_cloak_is_special_equipment_not_body_armour():
    corsair = find_profile("dark-elves", "corsairs")
    options = set(equipment_options_for_profile("dark-elves", "corsairs"))
    assert "Sea Dragon cloak" not in corsair.armors
    assert "Sea Dragon cloak" in options


def test_every_visible_weapon_and_armour_has_a_compact_tooltip():
    weapons = weapon_descriptions()
    armour = armour_descriptions()

    assert all(weapons.get(item) for item in WEAPONS_ALL)
    assert all(armour.get(item) for item in BODY_ARMORS)
    assert max(map(len, weapons.values())) <= 420
    assert max(map(len, armour.values())) <= 420


def test_canonical_candidate_catalog_is_complete():
    bands = load_bands()
    assert len(bands) == 49
    assert sum(len(band.profiles) for band in bands) == 318
    assassin = find_profile("skaven-clan-eshin", "assassin-adept")
    assert assassin.stats == {"HA": 4, "F": 4, "R": 3, "H": 1, "I": 5, "A": 1}
    assert {"Sword", "Dagger", "Spear"} <= set(assassin.weapons)
    assert "Light armour" in assassin.armors
    assert "Mighty Blow" in assassin.skills


def test_attack_replacement_skills_are_not_hidden_by_starting_attacks():
    duelist = find_profile("hochland-bandits", "duelist")
    assert duelist.stats["A"] == 1
    assert "Swordmaster" in duelist.combat_traits["starting_skills"]


def test_natural_profiles_receive_a_neutral_natural_weapon():
    rat_ogre = find_profile("skaven-clan-pestilens", "rat-ogre")
    assert rat_ogre.weapons == ("Natural attacks",)
    assert not rat_ogre.armors
    assert not rat_ogre.skills


def test_band_skills_are_contextual_and_profile_access_is_respected():
    bands = load_bands()
    skaven = next(band for band in bands if band.band_id == "skaven-clan-eshin")
    assert {"Black Hunger", "Art of Silent Death"} <= {
        skill.name for skill in skaven.skills
    }
    assassin = find_profile("skaven-clan-eshin", "assassin-adept")
    verminkin = find_profile("skaven-clan-eshin", "verminkin")
    assert "Black Hunger" in assassin.skills
    assert "Black Hunger" not in verminkin.skills


def test_skill_section_headings_are_not_exposed_as_selectable_skills():
    bands = {band.band_id: band for band in load_bands()}
    sisters = {skill.name for skill in bands["sisters-of-sigmar"].skills}
    ostlanders = {skill.name for skill in bands["ostlanders"].skills}
    pit_fighters = {skill.name for skill in bands["pit-fighters"].skills}

    assert "Special Skills" not in sisters
    assert "Ostlander Special Skills" not in ostlanders
    assert "Troll Slayer Special Skills" not in pit_fighters
    assert {
        "Sign of Sigmar", "Protection of Sigmar", "Utter Determination",
        "Righteous Fury", "Absolute Faith",
    } <= sisters


def test_amazon_special_skills_use_ui_title_case_without_the_section_heading():
    amazon = next(band for band in load_bands() if band.band_id == "amazons-lustria")
    names = {skill.name for skill in amazon.skills}

    assert "Amazon Special Skills" not in names
    assert names == {
        "Skink Hunter", "Elixir of Life", "Mesmerising Dance",
        "Savage Fury", "Concealment",
    }


def test_every_profile_exposes_exactly_the_skills_from_its_allowed_categories():
    for band in load_bands():
        for profile in band.profiles:
            flattened = tuple(
                skill
                for skills in profile.skills_by_category.values()
                for skill in skills
            )
            assert profile.skills == flattened, (band.name, profile.name)
            assert len(profile.skills) == len(set(profile.skills)), (
                band.name, profile.name,
            )


def test_skill_columns_are_preserved_for_mordheim_profiles():
    sorceress = find_profile("dark-elves", "dark-elf-sorceress")
    beastmaster = find_profile("dark-elves", "beastmaster")
    assert "academic" in sorceress.skills_by_category
    assert "combat" not in sorceress.skills_by_category
    assert "combat" in beastmaster.skills_by_category
    assert "academic" not in beastmaster.skills_by_category


def test_candidate_workbook_round_trip_uses_only_the_current_format(tmp_path: Path):
    path = tmp_path / "candidate.xlsx"
    payload = _payload()
    payload["config"].update({
        "has_sea_dragon_cloak": True,
        "preparations": ["Crimson Shade", "Mandrake Root"],
        "main_poison": "Black Lotus",
    })
    payload["house_rules"] = {
        "anti_dual": True, "hard_armour": True, "cheap_armour": False,
    }
    save_candidate_workbook(path, payload)
    restored = load_candidate_workbook(path)
    assert restored["config"]["candidate_name"] == "Rat with papers"
    assert restored["house_rules"]["anti_dual"] is True
    assert [profile["enemy_name"] for profile in restored["enemies"]["profiles"]] == [
        "Brute", "Fast"
    ]

    workbook = load_workbook(path)
    assert workbook.sheetnames[:3] == [SUMMARY_SHEET, ENEMIES_SHEET, RESULTS_INDEX_SHEET]
    assert workbook[DATA_SHEET].sheet_state == "veryHidden"
    assert workbook[SUMMARY_SHEET]["A1"].value == "MORDHEIM · SIMULATION PROFILE"
    assert workbook[SUMMARY_SHEET].auto_filter.ref is None
    enemy_values = [
        cell.value for row in workbook[ENEMIES_SHEET].iter_rows() for cell in row
    ]
    assert "ENEMY 1 · Brute" in enemy_values
    assert "ENEMY 2 · Fast" in enemy_values
    summary_values = [
        cell.value for row in workbook[SUMMARY_SHEET].iter_rows() for cell in row
    ]
    assert "QUICK REFERENCE" in summary_values
    assert "Skill: Mighty Blow" in summary_values
    assert "ACTIVE HOUSE RULES" in summary_values
    assert "Two-weapon Penalty" in summary_values
    assert "Harder Armour" in summary_values
    assert "Cheaper Armour" not in summary_values
    assert any(
        value and all(item in str(value) for item in (
            "Sea Dragon cloak", "Crimson Shade", "Mandrake Root",
            "Black Lotus",
        ))
        for value in summary_values
    )
    changed = _payload()
    changed["config"]["HA"] = 5
    save_candidate_workbook(path, changed)
    workbook = load_workbook(path, data_only=False)
    assert load_candidate_workbook(path)["config"]["HA"] == 5

    workbook[DATA_SHEET]["A1"] = "MORDHEIM_WORKBOOK_V0"
    workbook.save(path)
    with pytest.raises(CandidateWorkbookError, match="not compatible"):
        load_candidate_workbook(path)


def test_workbook_saves_and_loads_each_simulation_on_its_own_sheet(tmp_path: Path):
    path = tmp_path / "simulations.xlsx"
    payload = _payload()
    payload["results"] = [
        {
            "format_version": FORMAT_VERSION,
            "target": "combos", "title": "Improvements",
            "generated_at": "2026-08-23T10:30:00", "iterations": 25000,
            "opponent": "Configurable opponent: Brute", "view": "optimal",
            "headers": ["Improvement", "free hand %", "Hit %"],
            "rows": [["BASELINE", 41.25, 0.0], ["Mighty Blow", 49.5, 8.25]],
            "table_data": {"Single": [["BASELINE", 41.25, 0.0], []]},
            "card_data": [{"Single": 41.25}, "Single", {"Single": "Sword"}],
        },
        {
            "format_version": FORMAT_VERSION,
            "target": "weapons", "title": "Weapon configurations",
            "generated_at": "2026-08-23T10:35:00", "iterations": 10000,
            "opponent": "Configurable opponent: Brute", "view": "equipment",
            "headers": ["Main", "Off hand", "Win %"],
            "rows": [["Sword", "Dagger", 52.75]],
            "table_data": {"Single": [40.0, []]}, "card_data": None,
        },
    ]

    save_candidate_workbook(path, payload)
    restored = load_candidate_workbook(path)
    assert [result["target"] for result in restored["results"]] == ["combos", "weapons"]

    workbook = load_workbook(path, data_only=False)
    result_sheets = [name for name in workbook.sheetnames if name.startswith(RESULT_SHEET_PREFIX)]
    assert len(result_sheets) == 2
    assert workbook[RESULTS_INDEX_SHEET]["B7"].value == "Improvements"
    assert workbook[RESULTS_INDEX_SHEET]["C8"].value == 10000
    first_values = [cell.value for row in workbook[result_sheets[0]].iter_rows() for cell in row]
    assert "Mighty Blow" in first_values
    assert 49.5 in first_values


def test_result_export_allows_rows_that_do_not_apply_to_every_combat_mode():
    table_data = (
        {
            "Single": [40.0, [["Sword || None", 45.0, 5.0]]],
            "Shield": [42.0, []],
            "Dual": [43.0, []],
            "TwoHand": [44.0, []],
        },
        {"Single": "Sword", "Shield": "Sword + shield",
         "Dual": "Sword + dagger", "TwoHand": "Double-handed weapon"},
    )

    headers, rows = MordheimApp._result_export_rows("weapons", table_data)

    weapon_row = next(row for row in rows if row[0] == "Sword")
    assert len(weapon_row) == len(headers)
    assert weapon_row[2] == 45.0
    assert weapon_row[3:6] == (None, None, None)
    assert weapon_row[-2:] == ("Single", "Sword")
